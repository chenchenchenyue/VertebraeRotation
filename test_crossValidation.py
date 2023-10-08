import torch
import openpyxl
import os
from PIL import Image
import pandas as pd
import torchvision.transforms as transforms
from network_all import MLP
import argparse
from torch.utils.data import TensorDataset, DataLoader
from itertools import count
from sklearn.model_selection import KFold
from sklearn.model_selection import train_test_split
from dataset import MedicalDataset

cudax=3
numVal = 100.0


###计算向量loss
def DIYlossadd(outputs1,outputs2,outputs3):
    ####其次三个向量相互垂直
    loss = 0
    for output1,output2,output3 in zip(outputs1,outputs2,outputs3):
        value1 = torch.dot(output1,output2)
        value2 = torch.dot(output2,output3)
        value3 = torch.dot(output1,output3)
        loss += torch.pow((value1),2)+torch.pow((value2),2)+torch.pow((value3),2)
    return loss/outputs1.size()[0]
def getCosVal(y_preds,y_trues):
    loss = 0
    y_pred = y_preds[0]
    y_true = y_trues[0]
    dotVal = torch.dot(y_true,y_pred)
    fenMother = torch.linalg.norm(y_true)*torch.linalg.norm(y_pred)
    returnVal = dotVal/fenMother
   
    return returnVal
def dotest(countfold,modelstate_dictFold):
    parser = argparse.ArgumentParser()
    parser.add_argument('--image_size', type = tuple, default = (1024, 1024))
    parser.add_argument('--batch_size', type = int, default = 1)
    parser.add_argument('--data_dir', type = str, default = 'cocodata')
    parser.add_argument('--data_set', type = str, default = 'test')
    parser.add_argument('--device', type = str, default = "cuda:3" if torch.cuda.is_available() else "cpu")

    parser.add_argument('--num_classes', type = int, default = 2)
    parser.add_argument('--max_epoch', type = int, default = 100)
    parser.add_argument('--model_save_path', type = str, default = 'checkpoint')
    args = parser.parse_args()

    excel1 = openpyxl.load_workbook('predict.xlsx')
    table1 = excel1["Sheet1"]
    model = MLP()
    model.load_state_dict(torch.load('{}'.format(os.path.join(args.model_save_path,modelstate_dictFold)),map_location='cuda:3'))
    model=model.cuda(cudax)
    model.eval()
    ##################### 加载数据

    test_loader = DataLoader(MedicalDataset(countfold,"val",args), batch_size = args.batch_size, shuffle = True,
                                       collate_fn = lambda x: tuple(zip(*x))
                                       )
  
    
    with torch.no_grad():
       
        for step, sample in enumerate(test_loader):

            image_F1 = torch.stack(sample[0]).cuda(cudax)
            target_F1 = sample[1]
            ori_image_F1 = sample[2]
            image_F2 = torch.stack(sample[3]).cuda(cudax)
            target_F2 = sample[4]
            ori_image_F2 = sample[5]
            ######
            targetL3_1 = sample[6]
            targetL3_2 = sample[7]
            targetL3_3 = sample[8]
            targetL3_1 =torch.tensor([tuple(eval(target)) for target in targetL3_1]).cuda(cudax)
            targetL3_2 = torch.tensor([tuple(eval(target)) for target in targetL3_2]).cuda(cudax)
            targetL3_3 = torch.tensor([tuple(eval(target)) for target in targetL3_3]).cuda(cudax)
            ######
            targetL4_1 = sample[9]
            targetL4_2 = sample[10]
            targetL4_3 = sample[11]
            targetL4_1 =torch.tensor([tuple(eval(target)) for target in targetL4_1]).cuda(cudax)
            targetL4_2 = torch.tensor([tuple(eval(target)) for target in targetL4_2]).cuda(cudax)
            targetL4_3 = torch.tensor([tuple(eval(target)) for target in targetL4_3]).cuda(cudax)
            ######
            targetL5_1 = sample[12]
            targetL5_2 = sample[13]
            targetL5_3 = sample[14]
            targetL5_1 =torch.tensor([tuple(eval(target)) for target in targetL5_1]).cuda(cudax)
            targetL5_2 = torch.tensor([tuple(eval(target)) for target in targetL5_2]).cuda(cudax)
            targetL5_3 = torch.tensor([tuple(eval(target)) for target in targetL5_3]).cuda(cudax)

            imageName = sample[15]
            
            image_F1_list = list(image.to(args.device) for image in image_F1)
            targets1 = [{k : v.to(args.device) for k, v in t.items()} for t in target_F1]
            image_F2_list = list(image.to(args.device) for image in image_F2)
            targets2 = [{k : v.to(args.device) for k, v in t.items()} for t in target_F2]
            
 
        
            lossoutput1,lossoutput2,outputs11,outputs12,outputs13,outputs21,outputs22,outputs23,outputs31,outputs32,outputs33=model(image_F1, image_F2,image_F1_list, image_F2_list,targets1,targets2,typemode="test")
            # lossmask1 = sum(loss for loss in lossoutput1.values())
            # lossmask2 = sum(loss for loss in lossoutput2.values())
    


            print(outputs11)
            print(targetL3_1)
            loss11 = getCosVal(outputs11,targetL3_1)
            loss12 = getCosVal(outputs12,targetL3_2)
            loss13 = getCosVal(outputs13,targetL3_3)
            
            loss21 = getCosVal(outputs21,targetL4_1)
            loss22 = getCosVal(outputs22,targetL4_2)
            loss23 = getCosVal(outputs23,targetL4_3)

            loss31 = getCosVal(outputs31,targetL5_1)
            loss32 = getCosVal(outputs32,targetL5_2)
            loss33 = getCosVal(outputs33,targetL5_3)

         

        
            idx = step*3
            table1.cell(idx+2, 1).value = imageName[0]+"-L3"
            table1.cell(idx+2, 2).value = str(targetL3_1.cpu().numpy())
            table1.cell(idx+2, 3).value = str(outputs11.cpu().numpy()[0])
            table1.cell(idx+2, 4).value = str(loss11.cpu().numpy())
            table1.cell(idx+2, 5).value = str(targetL3_2.cpu().numpy())
            table1.cell(idx+2, 6).value = str(outputs12.cpu().numpy()[0])
            table1.cell(idx+2, 7).value = str(loss12.cpu().numpy())
            table1.cell(idx+2, 8).value = str(targetL3_3.cpu().numpy())
            table1.cell(idx+2, 9).value = str(outputs13.cpu().numpy()[0])
            table1.cell(idx+2, 10).value = str(loss13.cpu().numpy())

            table1.cell(idx+3, 1).value = imageName[0]+"-L4"
            table1.cell(idx+3, 2).value = str(targetL4_1.cpu().numpy())
            table1.cell(idx+3, 3).value = str(outputs21.cpu().numpy()[0])
            table1.cell(idx+3, 4).value = str(loss21.cpu().numpy())
            table1.cell(idx+3, 5).value = str(targetL4_2.cpu().numpy())
            table1.cell(idx+3, 6).value = str(outputs22.cpu().numpy()[0])
            table1.cell(idx+3, 7).value = str(loss22.cpu().numpy())
            table1.cell(idx+3, 8).value = str(targetL4_3.cpu().numpy())
            table1.cell(idx+3, 9).value = str(outputs23.cpu().numpy()[0])
            table1.cell(idx+3, 10).value = str(loss23.cpu().numpy())

            table1.cell(idx+4, 1).value = imageName[0]+"-L5"
            table1.cell(idx+4, 2).value = str(targetL5_1.cpu().numpy())
            table1.cell(idx+4, 3).value = str(outputs31.cpu().numpy()[0])
            table1.cell(idx+4, 4).value = str(loss31.cpu().numpy())
            table1.cell(idx+4, 5).value = str(targetL5_2.cpu().numpy())
            table1.cell(idx+4, 6).value = str(outputs32.cpu().numpy()[0])
            table1.cell(idx+4, 7).value = str(loss32.cpu().numpy())
            table1.cell(idx+4, 8).value = str(targetL5_3.cpu().numpy())
            table1.cell(idx+4, 9).value = str(outputs33.cpu().numpy()[0])
            table1.cell(idx+4, 10).value = str(loss33.cpu().numpy())
            

    excel1.save('predict_fold{}.xlsx'.format(countfold))


# import torch
# import cv2
# import argparse
# import numpy as np
# import random
# from pathlib import Path
# from torchvision.models.detection import maskrcnn_resnet50_fpn
# from torchvision.models.detection.faster_rcnn import FastRCNNPredictor
# from torch.utils.data import DataLoader
# from dataset import MedicalDataset
# class Solver :
#     def __init__(self, args):
#         self.args = args

#     def test(self):
#         self.args = args
#         # 制作训练集，要注意输出的数据状态，分割任务中因为每张图片的目标数量可能不一样，因此无法进行张量拼接
#         # 因此，对于图片或者标注信息我们要获取列表型数据，由参数 collate_fn 决定，可以查查该参数的用法
#         self.test_loader = DataLoader(MedicalDataset(args), batch_size = args.batch_size, shuffle = True,
#                                        collate_fn = lambda x: tuple(zip(*x)))
#         # 这部分与训练一样
#         model = maskrcnn_resnet50_fpn(pretrained = True)
#         in_features = model.roi_heads.box_predictor.cls_score.in_features
#         model.roi_heads.box_predictor = FastRCNNPredictor(in_features, num_classes = self.args.num_classes)
#         model = model.to(self.args.device)
#         model.eval()
#         # 加载已经保存的 fine-tuning 模型
#         ckpt = torch.load(Path(self.args.model_save_path).joinpath(f'mask_rcnn_with_newsssss{self.args.max_epoch}epochs.pth').__str__())
#         model.load_state_dict(ckpt)

#         # 读入图片数据并进行维度转换等操作
#         for images, targets,ori_image in self.test_loader :
#             images = list(image.to(self.args.device) for image in images)
#             targets = [{k : v.to(self.args.device) for k, v in t.items()} for t in targets]
#             # 损失，如果输入了 target 则输出损失，否则输出的是预测分数、框、分割等等信息
       

#             with torch.no_grad() :
#                 # 预测
#                 pred = model(images)
    
#             im = ori_image[0].astype(np.uint8)
#             im = cv2.cvtColor(im, cv2.COLOR_RGB2BGR)
#             im2 = im.copy()
#             showed = 0
#             for i in range(len(pred[0]['masks'])):
#                 # mask，[i, 0] 之所以要有 0，是因为 mask 输出维度为 [N, 1, H, W]
#                 msk = pred[0]['masks'][i, 0].detach().cpu().numpy()
#                 # 置信度
#                 showed = pred[0]['scores'][i].detach().cpu().numpy()
#                 # 置信分数超过 0.5 则接受该 mask
#                 if showed > 0.5:
#                     # 得到最终的分割掩膜，对软掩膜进行阈值处理，一般取值为 0.5
#                     im2[:, :, 0][msk > 0.5] = random.randint(0, 255)
#                     im2[:, :, 1][msk > 0.5] = random.randint(0, 255)
#                     im2[:, :, 2][msk > 0.5] = random.randint(0, 255)

#             cv2.imwrite(f'testImage/{str(showed)}.jpg', np.hstack([im, im2]))
#             # cv2.imshow(str(showed), np.hstack([im, im2]))
#             # cv2.waitKey()


if __name__ == '__main__':
    path = "./new裁剪图片/"
    dataSetdata = []
    dataSetlabel = []
    noSaveDatas = ['caoyue_LTorL', 'caoyue_LTorR', 'caoyue_UBenL', 'caoyue_UTorR', 'ChengGuangchuan_LFlex', 'ZhangJianxinyiyou_UExten', 'ZhangJianxinyiyou_UFlex','zhoupingxiangyiyou_UTorR']
    with open("angleoutput.txt","r",encoding="utf8") as fdataset:
        datas = fdataset.readlines()
        for data in datas:
            dataLine = data.strip().split("\t")
            if dataLine[0].split("-")[0] not in noSaveDatas:
                dataSetdata.append(dataLine[0].split("-")[0])


    dataSetdata = list(set(dataSetdata))
    print(len(dataSetdata))
    FOLD = 5
    sfolder = KFold(n_splits=FOLD,shuffle=True,random_state=24)
    tr_folds = []
    val_folds = []
    countFold = 1
    modelstate_dict = []
    for train_idx, val_idx in sfolder.split(dataSetdata):
        tr_folds.append(train_idx)
        val_folds.append(val_idx)
        print("--------------------test fold {}--------------------".format(countFold))
        dotest(countFold,modelstate_dict[countFold-1])
        countFold += 1




